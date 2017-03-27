import openpyxl
#from sys import argv
import jieba
from jieba import cut

compeny_word_lists = ['公司','馆',"法院","技术","学院","事务所","工会","分会","园","别墅","协会","医院",'集团','平台','出版社','委员会','官方','中心','大学','会所']

def begining():
    print('''这是一个公众号的处理脚本
    功能一：统计表格中公众号名称中特定关键字的数量，以查知企业/机构数量；
    功能二：将企业/机构和它们的代码，整理成为新的名单；
    功能三：分离企业号和个人号，可视作功能二的强化版本。
    ''')
    choose = input('''请选择输入……
    功能一：1
    功能二：2
    功能三：3
    ''')
    if choose == '1':
        ID_Count()
    elif choose == '2':
        ID_Cllection()
    elif choose == '3':
        ID_Separation()
    

def ID_Count():
    
    print ('''你好，这是一个统计公众号的脚本
    该脚本可以统计公众号的关键词
    简单推测有多少公号属于“企业/机构号”。
    推测的原理是"关键字"检索（关键词的匹配还在研究中）
    ***注意，请确保被统计的表单是xlsx格式***
    ***注意，请确保公众号名称在“B”列纵向排列***
    ***注意，请确保文件在以下地址：C:\\Users\\aorui\\Documents\\aorick\\my word***
    ***如果结果依然出错，应为统计数据中有"空数据"***
    ***表格的第一栏名称必须为"Sheet"***
    ''')

    filename = input('准备好了请输入文件名（必须.xlsx文件）：')
    wb = openpyxl.load_workbook(filename + '.xlsx')
    sheet = wb.get_sheet_by_name("Sheet")

    print('该表单最大列：%s' % sheet.max_column)
    print('该表单最大行：%s' % sheet.max_row)

    compeny_word_lists = ['司','国','馆','室','院','社','委','队','园','官','心','校','会','部']

    print("下面需要定位信息处理位置，列信息/行信息")
    the_column = input('输入要处理的列（A、B or C……):')
    start_row = input('输入起始行：')
    end_row = input('输入终点行：')

    start_spot = the_column + start_row
    end_spot = the_column + end_row
    
    n = 0
    for rowOfCellObj in sheet[start_spot:end_spot]:
        #print(rowOfCellObj)#这里是一个tuple (<Cell Sheet1.A97>, <Cell Sheet1.B97>)
        for CellObj in rowOfCellObj:
            #print(CellObj) #这里是一个属性 <Cell Sheet1.A7>
            for divc_words in str(CellObj.value):
                if divc_words in compeny_word_lists:
                    n += 1
                    break
                else:
                    pass
    print('''
    计算完毕……
    公众号数量：%s 
    企业号数量：%s
    本批样本企业号比例：%s%%
    可进入脚本修改字符库
    再见！'''% (int(end_row) - int(start_row) + 1,n,int(100 * n / (int(end_row) - int(start_row)))))

def Creat_Table(Table_Name,Sheet_Num=0):
    #Table_Name = input('请输入需要建立的Excel文档名称：')
    wb = openpyxl.Workbook()
    #sheet = wb.create_sheet(title='Sheet1')
    #关于修改Sheet名称的问题我还不熟悉，因此统一修改成为Sheet
    sheet = wb.get_active_sheet()
    for x in range(Sheet_Num):
        Sheet_Name = 'Sheet'+ 'x'
        Sheet_Name = wb.create_sheet()

    wb.save(Table_Name + '.xlsx')
    print("""
    
    """)
    print('文档%s.xlsx 已经建立……内含工作表%s' % (Table_Name, wb.sheetnames))

def ID_Cllection(start_spot = None ,end_spot = None):
    print ('''
    功能二：将企业/机构和它们的代码，整理成为新的名单；
    ***注意，请确保被统计的表单是xlsx格式***
    ***注意，请确保公众号代码在“A”列，名称在“B”列纵向排列***
    ***注意，请确保文件在以下地址：C:\\Users\\aorui\\Documents\\aorick\\my word***
    ***如果结果依然出错，应为统计数据中有‘空数据’***
    ''')

    filename = input('准备好了请输入素材文件名称（必须.xlsx文件）：')
    wb = openpyxl.load_workbook(filename + '.xlsx')
    sheet = wb.get_sheet_by_name("Sheet")

    print('该表单最大列：%s' % sheet.max_column)
    print('该表单最大行：%s' % sheet.max_row)

    #compeny_word_lists = ['海']

    print("下面需要定位信息处理位置，列信息/行信息")
    #the_column = input('输入要处理的列（A、B or C……):')
    start_row = input('输入起始行：')
    end_row = input('输入终点行：')

    start_spot = 'A' + str(start_row)
    end_spot = 'B' + str(end_row)

    print('首先，我们需要建立一个空白表格……')
    Table_Name = input('请输入空白的Excel文档名称：')
    Creat_Table(Table_Name)
    print('\n正在工作，将结果写入%s.xlsx……' %Table_Name)
    n = 0
    Nwb = openpyxl.load_workbook(Table_Name + '.xlsx')
    Nsheet = Nwb.get_sheet_by_name('Sheet')

    for rowOfCellObj in sheet[start_spot:end_spot]:
        #print(rowOfCellObj) #这里是一个tuple
        #for CellObj in rowOfCellObj:
            #print(CellObj) 这里是一个属性
        for divc_words in jieba.cut(rowOfCellObj[1].value, cut_all=False):
            if divc_words in compeny_word_lists:
                n += 1
                Nsheet['B'+ str(n)] = rowOfCellObj[1].value
                Nsheet['A'+ str(n)] = rowOfCellObj[0].value
                break
            else:
                print(rowOfCellObj[1].value)
                break
    Nwb.save(Table_Name + '.xlsx')
    print('数据写入完毕')

def ID_Separation():
    print ('''
    功能三：这是功能二的加强版
    将企业/机构和它们的代码整理成为新的Sheet；
    将流量主和它们的代码整理成为另一个Sheet。
    ***注意，请确保被统计的表单是xlsx格式***
    ***注意，请确保公众号代码在“A”列，名称在“B”列纵向排列***
    ***注意，请确保文件在以下地址：C:\\Users\\aorui\\Documents\\aorick\\my word***
    ***如果结果依然出错，应为统计数据中有‘空数据’***
    ''')
    filename = input('准备好了请输入素材文件名称（必须.xlsx文件）：')
    wb = openpyxl.load_workbook(filename + '.xlsx')
    sheet = wb.get_sheet_by_name("Sheet")
    #print(sheet)
    print('该表单最大列：%s' % sheet.max_column)
    print('该表单最大行：%s' % sheet.max_row)

    print("下面需要定位信息处理位置，列信息/行信息")
    the_column = input('输入要处理的列（A、B or C……):')
    start_row = input('输入起始行：')
    end_row = input('输入终点行：')

    start_spot = 'A' + str(start_row)
    end_spot = 'B' + str(end_row)

    print('首先，我们需要建立一个空白表格……')
    Table_Name = input('请输入空白的Excel文档名称：')
    Creat_Table(Table_Name,1)
    print('''\n正在工作，将结果写入%s.xlsx……
    Sheet：企业名录
    Sheet1：流量主名录
    ''' %Table_Name)
    n = 0
    m = 0
    Nwb = openpyxl.load_workbook(Table_Name + '.xlsx')
    Sheet_of_Companies = Nwb.get_sheet_by_name('Sheet')
    Sheet_of_Users = Nwb.get_sheet_by_name('Sheet1')
    #print(wb.sheetnames)

    for rowOfCellObj in sheet[start_spot: end_spot]:
        #print(rowOfCellObj) #这里是一个tuple
        #for CellObj in rowOfCellObj:
        #print(CellObj) 这里是一个属性
        div_words = jieba.cut_for_search(rowOfCellObj[1].value)
        jiao_ji = set(div_words).intersection(compeny_word_lists)
        if jiao_ji == set():
            m += 1
            Sheet_of_Users['B'+ str(m)], Sheet_of_Users['A'+ str(m)] = rowOfCellObj[1].value,rowOfCellObj[0].value
            
        else:
            print(rowOfCellObj[1].value)
            n += 1
            Sheet_of_Companies['B'+ str(n)], Sheet_of_Companies['A'+ str(n)] = rowOfCellObj[1].value,rowOfCellObj[0].value
    
    print("执行完毕，总计录入流量主%s个,写入企业主%s" % (m,n))
    #wb.save()
    #Nwb.save(Table_Name + '.xlsx')
    Nwb.save(Table_Name + '.xlsx')

#ID_Count(start_spot, end_spot)
begining()
#Creat_Table()
#ID_Count_2('A1', 'B100')
#ID_Cllection()
#ID_Separation()
#make_company_list("text01")